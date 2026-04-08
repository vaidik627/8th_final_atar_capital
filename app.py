"""
CIM Financial Intelligence — Flask Server
Atar Capital · M&A Deal Analysis Platform
"""

import os
import sys
import json
import uuid
import logging
import tempfile
import threading
from pathlib import Path
from flask import Flask, render_template, request, jsonify, send_file, make_response

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from cim_extractor import CIMParser, LLMExtractor, _load_env, DEFAULT_CLIENT_REQ

_load_env()

app = Flask(__name__, template_folder="templates", static_folder="static")
app.config["MAX_CONTENT_LENGTH"] = 200 * 1024 * 1024
app.config["JSON_SORT_KEYS"] = False  # preserve extraction field order in API responses

jobs: dict = {}
jobs_lock = threading.Lock()

PROVIDER_ENV = {
    "nvidia":    "NVIDIA_API_KEY",
    "anthropic": "ANTHROPIC_API_KEY",
    "openai":    "OPENAI_API_KEY",
    "ollama":    "OLLAMA_API_KEY",
}

DEFAULT_REQ = DEFAULT_CLIENT_REQ


class _LogHandler(logging.Handler):
    def __init__(self, job_id):
        super().__init__()
        self.job_id = job_id

    def emit(self, record):
        with jobs_lock:
            if self.job_id in jobs:
                jobs[self.job_id]["log"].append(self.format(record))


def _run_extraction(job_id, pdf_path, provider, model, api_key, deal_value=None):
    handler = _LogHandler(job_id)
    handler.setFormatter(logging.Formatter("%(levelname)s — %(message)s"))
    logging.getLogger().addHandler(handler)

    def _set(status, **kw):
        with jobs_lock:
            if job_id in jobs:
                jobs[job_id]["status"] = status
                jobs[job_id].update(kw)

    try:
        _set("parsing")
        cim = CIMParser(pdf_path)
        cim.parse_pdf()

        _set("filtering")
        relevant = cim.find_financial_sections() or list(cim.extracted_tables)

        _set("extracting")
        extractor = LLMExtractor(api_key=api_key, provider=provider, model_name=model)
        raw = extractor.extract_fields(relevant, client_requirements=DEFAULT_REQ)
        if raw is None:
            raise RuntimeError("LLM returned no data. Check API key and model.")

        # Generate investment recommendation if deal value provided
        if deal_value:
            try:
                dv_float = float(str(deal_value).replace(',', '').replace('$', '').strip())
                _set("recommending")
                recommendation = extractor.generate_investment_recommendation(raw, dv_float)
                if recommendation:
                    ordered = {}
                    for key in raw:
                        if key == "Company_Summary":
                            ordered["Investment_Recommendation"] = recommendation
                        ordered[key] = raw[key]
                    if "Investment_Recommendation" not in ordered:
                        ordered["Investment_Recommendation"] = recommendation
                    raw = ordered
            except Exception as e:
                logging.warning(f"Investment recommendation skipped: {e}")

        out_dir = Path("extracted_results")
        out_dir.mkdir(exist_ok=True)
        base = Path(pdf_path).stem.replace(" ", "_")
        raw_path = str(out_dir / f"{base}_extracted.json")
        Path(raw_path).write_text(json.dumps(raw, indent=4))

        _set("done", raw=raw, raw_path=raw_path)

    except Exception as exc:
        logging.error(f"Extraction error: {exc}")
        _set("error", error=str(exc))
    finally:
        logging.getLogger().removeHandler(handler)
        try:
            os.unlink(pdf_path)
        except OSError:
            pass


@app.route("/")
def index():
    resp = make_response(render_template("index.html"))
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    return resp


@app.route("/results")
def results():
    resp = make_response(render_template("results.html"))
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    return resp


@app.route("/api/extract", methods=["POST"])
def api_extract():
    pdf_file = request.files.get("pdf")
    provider  = request.form.get("provider", "nvidia")
    model     = request.form.get("model", "meta/llama-3.3-70b-instruct")
    api_key   = request.form.get("api_key", "").strip()
    # deal_name and deal_value are metadata only — stored in job for display
    deal_name  = request.form.get("deal_name", "").strip()
    deal_value = request.form.get("deal_value", "").strip()

    if not pdf_file or not pdf_file.filename.lower().endswith(".pdf"):
        return jsonify({"error": "Please upload a valid PDF file."}), 400
    if not api_key:
        api_key = os.getenv(PROVIDER_ENV.get(provider, ""), "")
    if not api_key:
        return jsonify({"error": f"API key missing for provider '{provider}'."}), 400

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    pdf_file.save(tmp.name)
    tmp.close()

    job_id = str(uuid.uuid4())
    with jobs_lock:
        jobs[job_id] = {
            "status": "queued", "log": [], "raw": None, "error": None,
            "deal_name": deal_name, "deal_value": deal_value,
            "filename": pdf_file.filename,
        }

    threading.Thread(
        target=_run_extraction,
        args=(job_id, tmp.name, provider, model, api_key, deal_value),
        daemon=True,
    ).start()
    return jsonify({"job_id": job_id}), 202


@app.route("/api/status/<job_id>")
def api_status(job_id):
    with jobs_lock:
        job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    return jsonify({
        "status": job["status"],
        "log":    job["log"][-20:],
        "error":  job.get("error"),
    })


@app.route("/api/result/<job_id>")
def api_result(job_id):
    with jobs_lock:
        job = jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    if job["status"] != "done":
        return jsonify({"error": "Not finished"}), 400
    return jsonify({
        "raw":        job["raw"],
        "deal_name":  job["deal_name"],
        "deal_value": job["deal_value"],
        "filename":   job["filename"],
    })


@app.route("/api/download/<job_id>")
def api_download(job_id):
    with jobs_lock:
        job = jobs.get(job_id)
    if not job or job["status"] != "done":
        return jsonify({"error": "Result not ready"}), 400
    path = job.get("raw_path")
    if path and os.path.exists(path):
        return send_file(path, as_attachment=True)
    import io
    return send_file(
        io.BytesIO(json.dumps(job["raw"], indent=4).encode()),
        mimetype="application/json",
        as_attachment=True,
        download_name="extracted.json",
    )


@app.route("/api/download-excel/<job_id>")
def api_download_excel(job_id):
    import io
    import openpyxl

    with jobs_lock:
        job = jobs.get(job_id)
    if not job or job["status"] != "done":
        return jsonify({"error": "Result not ready"}), 400

    raw = job.get("raw") or {}

    template_path = Path(__file__).parent / "Prebid V31  Template (1).xlsx"
    if not template_path.exists():
        return jsonify({"error": "Excel template not found"}), 500

    # Copy template to in-memory buffer (preserves formulas)
    buf = io.BytesIO()
    with open(template_path, "rb") as f:
        buf.write(f.read())
    buf.seek(0)

    wb = openpyxl.load_workbook(buf)
    ws = wb["Sheet1"]

    # ── Column mapping helpers ──────────────────────────────────────────────
    # Excel columns: I=FY19, J=FY20, K=FY21 (last 3 actuals)
    #                L=Year1, M=Year2, N=Year3, O=Year4, P=Year5 (projections)
    COL_HIST = ["I", "J", "K"]   # last 3 actuals
    COL_PROJ = ["L", "M", "N", "O", "P"]  # up to 5 projected

    def _sorted_keys(d, suffix):
        """Return year keys ending in suffix, sorted ascending by year."""
        if not isinstance(d, dict):
            return []
        return sorted(
            [k for k in d if k.endswith(suffix) and d[k] is not None],
            key=lambda k: k[:4]
        )

    # ── Build canonical year column lists ────────────────────────────────────
    # Always produce exactly 3 hist + 5 proj slots.
    # Missing projected slots → extrapolate year labels and fill values with
    # the average of the 3 historical values for each field.

    def _year_label(key):
        """Convert '2022_A' → '2022A', '2025_E' → '2025E'"""
        return key.replace("_A", "A").replace("_E", "E")

    ref_field = raw.get("Total_Revenue") or raw.get("Adj_EBITDA") or {}
    ref_field = ref_field if isinstance(ref_field, dict) else {}

    # Actual year keys (last 3)
    all_actuals = _sorted_keys(ref_field, "_A")
    hist_keys = all_actuals[-len(COL_HIST):]  # up to 3

    # Projected year keys (up to 5 from data)
    all_proj = _sorted_keys(ref_field, "_E")
    proj_keys = all_proj[:len(COL_PROJ)]  # up to 5

    # How many synthetic projected years do we need?
    missing_proj = len(COL_PROJ) - len(proj_keys)

    # Extrapolate synthetic year labels beyond last known projected year
    if missing_proj > 0:
        if proj_keys:
            last_proj_year = int(proj_keys[-1][:4])
        elif hist_keys:
            last_proj_year = int(hist_keys[-1][:4])
        else:
            last_proj_year = 0
        synthetic_labels = [f"{last_proj_year + j + 1}E" for j in range(missing_proj)]
    else:
        synthetic_labels = []

    # Full projected label list (real + synthetic)
    full_proj_labels = [_year_label(k) for k in proj_keys] + synthetic_labels

    # ── Row 4 — Year headers (I4:P4) ─────────────────────────────────────────
    for i, yk in enumerate(hist_keys):
        col = COL_HIST[len(COL_HIST) - len(hist_keys) + i]
        ws[f"{col}4"] = _year_label(yk)

    for i, lbl in enumerate(full_proj_labels):
        ws[f"{COL_PROJ[i]}4"] = lbl

    # ── Core write helpers ────────────────────────────────────────────────────

    def _write_field(field_key, row, hist_cols=COL_HIST, proj_cols=COL_PROJ, force_negative=False):
        """
        3-case logic (applied per-field independently):
          Case 1 — hist + proj both present → fill directly
          Case 2 — hist present, proj all absent → fill hist + avg of hist fills all proj slots
          Case 3 — both absent → write nothing (leave blank)
        force_negative: if True, ensure all written values are negative (CAPEX).
        """
        d = raw.get(field_key)
        if not isinstance(d, dict):
            return  # Case 3 — field not extracted at all

        actuals  = _sorted_keys(d, "_A")[-len(hist_cols):]
        proj     = _sorted_keys(d, "_E")[:len(proj_cols)]

        hist_has = len(actuals) > 0
        proj_has = len(proj) > 0

        if not hist_has and not proj_has:
            return  # Case 3 — all null

        def _safe_val(v):
            if v is None:
                return None
            if force_negative:
                return -abs(v) if v != 0 else 0
            return v

        # Write historical
        for i, yk in enumerate(actuals):
            col = hist_cols[len(hist_cols) - len(actuals) + i]
            ws[f"{col}{row}"] = _safe_val(d[yk])

        if proj_has:
            # Case 1 — write projected directly
            for i, yk in enumerate(proj):
                val = _safe_val(d.get(yk))
                if val is not None:
                    ws[f"{proj_cols[i]}{row}"] = val
        else:
            # Case 2 — proj absent: fill all proj slots with hist average
            if hist_has:
                hist_vals = [d[k] for k in actuals if d.get(k) is not None]
                if hist_vals:
                    avg = round(sum(hist_vals) / len(hist_vals))
                    avg = _safe_val(avg)
                    for i in range(len(proj_cols)):
                        ws[f"{proj_cols[i]}{row}"] = avg

    def _write_field_hist_only(field_key, row, hist_cols=COL_HIST):
        """
        Historical columns only — 3 cases on hist; never touch proj columns.
        Used for: Onex_Adjustments (proj is a formula).
        """
        d = raw.get(field_key)
        if not isinstance(d, dict):
            return
        actuals = _sorted_keys(d, "_A")[-len(hist_cols):]
        if not actuals:
            return
        for i, yk in enumerate(actuals):
            col = hist_cols[len(hist_cols) - len(actuals) + i]
            ws[f"{col}{row}"] = d[yk]

    def _write_field_formula_proj(field_key, row, hist_cols=COL_HIST, proj_cols=COL_PROJ):
        """
        Historical: 3-case logic (hist present → write; both absent → blank).
        Projected: write ONLY if extracted — never apply avg fill (formula handles it).
        Used for: Interest_Expense, WC_Change.
        """
        d = raw.get(field_key)
        if not isinstance(d, dict):
            return

        actuals  = _sorted_keys(d, "_A")[-len(hist_cols):]
        proj     = _sorted_keys(d, "_E")[:len(proj_cols)]

        hist_has = len(actuals) > 0
        proj_has = len(proj) > 0

        if not hist_has and not proj_has:
            return  # both absent → leave blank

        # Write historical (3-case: Case 2 avg fills hist slots if some null, but all absent = skip)
        if hist_has:
            for i, yk in enumerate(actuals):
                col = hist_cols[len(hist_cols) - len(actuals) + i]
                ws[f"{col}{row}"] = d[yk]

        # Write projected ONLY if extracted — no avg fill, formula handles missing
        if proj_has:
            for i, yk in enumerate(proj):
                val = d.get(yk)
                if val is not None:
                    ws[f"{proj_cols[i]}{row}"] = val

    # ── Write input cells ─────────────────────────────────────────────────────

    # B1 — Company / Deal name
    deal_name_val = job.get("deal_name", "").strip()
    if deal_name_val:
        ws["B1"] = deal_name_val

    # Clear hardcoded template values in SG&A projection columns (L13:P13)
    for col in COL_PROJ:
        ws[f"{col}13"] = None

    # Row 7  — Net Revenue
    _write_field("Total_Revenue", 7)

    # Row 10 — Gross Margin
    _write_field("Gross_Margin", 10)

    # Row 13 — SG&A
    _write_field("SGA", 13)

    # Row 16 — 1x Adjustments (historical only — proj is formula)
    _write_field_hist_only("Onex_Adjustments", 16)

    # Row 21 — Other Expense / (Income)
    _write_field("Other_Expense", 21)

    # Row 22 — Interest Expense (hist: 3-case; proj: write if extracted, else formula)
    _write_field_formula_proj("Interest_Expense", 22)

    # Row 23 — Depreciation
    _write_field("Depreciation", 23)

    # Row 32 — CAPEX (hist + proj, always negative, 3-case)
    _write_field("CAPEX", 32, force_negative=True)

    # Row 33 — Working Capital Change (hist: 3-case; proj: write if extracted, else formula)
    _write_field_formula_proj("WC_Change", 33)

    # C7 — Accounts Receivable (single scalar)
    ar = raw.get("AR")
    if ar is not None and not isinstance(ar, dict):
        ws["C7"] = ar

    # C8 — Inventory (single scalar)
    inv = raw.get("Inventory")
    if inv is not None and not isinstance(inv, dict):
        ws["C8"] = inv

    # C9 — Machinery & Equipment (single scalar)
    me = raw.get("ME_Equipment")
    if me is not None and not isinstance(me, dict):
        ws["C9"] = me

    # C11 — Building & Land (single scalar)
    bl = raw.get("Building_Land")
    if bl is not None and not isinstance(bl, dict):
        ws["C11"] = bl

    # ── Serialize and return ────────────────────────────────────────────────
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    deal_name = job.get("deal_name") or "CIM"
    safe_name = "".join(c for c in deal_name if c.isalnum() or c in " _-").strip() or "CIM"
    download_name = f"{safe_name}_Prebid.xlsx"

    return send_file(
        out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=download_name,
    )


@app.route("/api/env-keys")
def api_env_keys():
    return jsonify({p: bool(os.getenv(v, "")) for p, v in PROVIDER_ENV.items()})


if __name__ == "__main__":
    import argparse as _ap
    _parser = _ap.ArgumentParser()
    _parser.add_argument("--port", type=int, default=5000)
    _args = _parser.parse_args()
    logging.basicConfig(level=logging.INFO, format="%(asctime)s — %(levelname)s — %(message)s")
    print(f"\n  CIM Intelligence · Atar Capital")
    print(f"  http://localhost:{_args.port}\n")
    app.run(debug=False, host="0.0.0.0", port=_args.port, threaded=True)
